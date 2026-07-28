"""Testes dos exportadores da execução (seção 16)."""

from __future__ import annotations

import csv
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from waypoint_etl.domain.entities.customer import Customer
from waypoint_etl.domain.entities.invoice import Invoice
from waypoint_etl.domain.enums.document_type import DocumentType
from waypoint_etl.domain.enums.entity_type import EntityType
from waypoint_etl.domain.enums.invoice_status import InvoiceStatus
from waypoint_etl.domain.value_objects.issue import error
from waypoint_etl.infrastructure.reports.exporters import (
    export_accepted,
    export_duplicates,
    export_rejected,
    run_output_dir,
)
from waypoint_etl.pipeline.deduplication.detector import (
    DeduplicationResult,
    DuplicateMatch,
)
from waypoint_etl.pipeline.validators.result import ValidatedRecord

VALID_CPF = "39053344705"


def _accepted(row: int = 2) -> ValidatedRecord:
    return ValidatedRecord(
        row_number=row,
        entity_type=EntityType.CUSTOMERS,
        values={"full_name": "Ana Silva", "document": VALID_CPF},
        entity=Customer(
            full_name="Ana Silva",
            document=VALID_CPF,
            document_type=DocumentType.CPF,
            external_id="ERP-1",
            email="ana@exemplo.com.br",
            created_at=datetime(2024, 3, 15),
        ),
    )


def _rejected(row: int = 3) -> ValidatedRecord:
    return ValidatedRecord(
        row_number=row,
        entity_type=EntityType.CUSTOMERS,
        values={"full_name": "Bruno", "document": "111", "email": "x"},
        issues=(
            error("invalid_document", "Documento inválido.", field="document"),
            error("invalid_email", "E-mail inválido.", field="email"),
        ),
    )


def test_run_output_dir_uses_the_run_id() -> None:
    assert run_output_dir(Path("exports"), "abc") == Path("exports/abc")


def test_accepted_csv_has_the_canonical_columns(tmp_path: Path) -> None:
    path = export_accepted(
        tmp_path / "accepted.csv", [_accepted()], EntityType.CUSTOMERS
    )

    with path.open(encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))

    assert list(rows[0]) == [
        "external_id",
        "full_name",
        "document",
        "email",
        "phone",
        "postal_code",
        "city",
        "state",
        "created_at",
    ]
    assert rows[0]["document"] == VALID_CPF
    assert rows[0]["created_at"] == "2024-03-15 00:00:00"
    assert rows[0]["phone"] == ""


def test_accepted_csv_keeps_the_real_document(tmp_path: Path) -> None:
    """O accepted é a carga de importação: mascarar aqui quebraria o destino."""
    path = export_accepted(
        tmp_path / "accepted.csv", [_accepted()], EntityType.CUSTOMERS
    )

    assert VALID_CPF in path.read_text(encoding="utf-8")


def test_accepted_csv_serializes_money_without_float(tmp_path: Path) -> None:
    record = ValidatedRecord(
        row_number=2,
        entity_type=EntityType.INVOICES,
        values={},
        entity=Invoice(
            external_id="NF-1",
            customer_document=VALID_CPF,
            issued_at=date(2024, 3, 1),
            due_at=date(2024, 3, 31),
            amount=Decimal("1234.50"),
            status=InvoiceStatus.OPEN,
        ),
    )

    path = export_accepted(
        tmp_path / "accepted.csv", [record], EntityType.INVOICES
    )

    with path.open(encoding="utf-8", newline="") as handle:
        row = next(csv.DictReader(handle))

    assert row["amount"] == "1234.50"
    assert row["status"] == "open"
    assert row["issued_at"] == "2024-03-01"


def test_accepted_csv_is_empty_when_nothing_is_valid(tmp_path: Path) -> None:
    path = export_accepted(tmp_path / "accepted.csv", [], EntityType.CUSTOMERS)

    lines = path.read_text(encoding="utf-8").strip().splitlines()

    assert len(lines) == 1  # apenas o cabeçalho


def test_rejected_xlsx_has_one_row_per_issue(tmp_path: Path) -> None:
    """Uma linha por problema: o usuário corrige campo a campo."""
    path = export_rejected(tmp_path / "rejected.xlsx", [_rejected()])

    workbook = load_workbook(path)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    assert rows[0][:6] == (
        "linha",
        "aba",
        "campo",
        "codigo",
        "severidade",
        "mensagem",
    )
    assert len(rows) == 3  # cabeçalho + duas issues
    assert rows[1][0] == 3
    assert rows[1][3] == "invalid_document"


def test_rejected_xlsx_keeps_the_source_values(tmp_path: Path) -> None:
    """Sem os valores de origem, o usuário não consegue corrigir a planilha."""
    path = export_rejected(tmp_path / "rejected.xlsx", [_rejected()])

    workbook = load_workbook(path)
    rows = list(workbook.active.iter_rows(values_only=True))
    workbook.close()

    assert "full_name" in rows[0]
    assert "Bruno" in rows[1]


def test_rejected_xlsx_is_valid_when_there_are_no_rejections(
    tmp_path: Path,
) -> None:
    path = export_rejected(tmp_path / "rejected.xlsx", [])

    workbook = load_workbook(path)
    rows = list(workbook.active.iter_rows(values_only=True))
    workbook.close()

    assert len(rows) == 1


def test_duplicates_csv_separates_exact_from_possible(tmp_path: Path) -> None:
    duplicates = DeduplicationResult(
        exact=(
            DuplicateMatch(
                row_number=5,
                matched_row_number=2,
                key="document",
                value=VALID_CPF,
                exact=True,
            ),
        ),
        possible=(
            DuplicateMatch(
                row_number=7,
                matched_row_number=3,
                key="phone",
                value="11987654321",
                exact=False,
                similarity=0.912,
            ),
        ),
    )

    path = export_duplicates(tmp_path / "duplicates.csv", duplicates)

    with path.open(encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))

    assert rows[0]["tipo"] == "exata"
    assert rows[0]["linha"] == "5"
    assert rows[0]["similaridade"] == ""
    assert rows[1]["tipo"] == "possivel"
    assert rows[1]["similaridade"] == "0.912"


def test_duplicates_csv_is_created_even_when_empty(tmp_path: Path) -> None:
    path = export_duplicates(tmp_path / "duplicates.csv", DeduplicationResult())

    assert path.read_text(encoding="utf-8").strip().splitlines() == [
        "linha,linha_correspondente,tipo,chave,valor,similaridade"
    ]


def test_exporters_create_the_directory(tmp_path: Path) -> None:
    target = tmp_path / "exports" / "run-1" / "accepted.csv"

    export_accepted(target, [_accepted()], EntityType.CUSTOMERS)

    assert target.is_file()
