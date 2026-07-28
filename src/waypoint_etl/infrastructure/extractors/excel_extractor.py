"""Extrator de planilhas Excel (``.xlsx``/``.xlsm``).

Usa openpyxl em modo somente leitura com ``data_only=True``: lê o último valor
calculado das fórmulas e nunca executa macros (seção 18).
"""

from __future__ import annotations

import zipfile
from collections.abc import Iterator
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from ...application.dto.extraction import (
    ExtractionOptions,
    ExtractionResult,
    SourceRecord,
)
from ...domain.enums.source_format import SourceFormat
from ...domain.errors import EmptySourceError, ExtractionError, SheetNotFoundError
from .base import (
    build_row_mapping,
    coerce_cell,
    ensure_readable_file,
    is_blank_row,
    normalize_headers,
)

SUPPORTED_SUFFIXES = frozenset({".xlsx", ".xlsm"})


class ExcelExtractor:
    """Lê planilhas Excel com suporte a abas e cabeçalho configurável."""

    @property
    def source_format(self) -> SourceFormat:
        return SourceFormat.EXCEL

    def supports(self, path: Path) -> bool:
        return path.suffix.lower() in SUPPORTED_SUFFIXES

    def extract(
        self, path: Path, options: ExtractionOptions | None = None
    ) -> ExtractionResult:
        opts = options or ExtractionOptions()
        ensure_readable_file(path)
        if opts.header_row < 1:
            raise ExtractionError(
                f"header_row deve ser maior ou igual a 1 (recebido: {opts.header_row})."
            )

        workbook = _open_workbook(path)
        try:
            sheet_names = tuple(workbook.sheetnames)
            worksheet = _select_sheet(workbook, sheet_names, opts.sheet, path)
            sheet_title = str(worksheet.title)
            rows = _iter_rows(worksheet)
            header, warnings = _read_header(rows, opts.header_row, path, sheet_title)
            records = tuple(_read_records(rows, header, sheet_title))
        finally:
            workbook.close()

        return ExtractionResult(
            source_name=path.name,
            source_format=SourceFormat.EXCEL,
            columns=header,
            records=records,
            sheet=sheet_title,
            available_sheets=sheet_names,
            warnings=warnings,
        )


def _open_workbook(path: Path) -> Any:
    """Abre a planilha em modo somente leitura, sem avaliar fórmulas."""
    try:
        return load_workbook(path, read_only=True, data_only=True, keep_links=False)
    except (InvalidFileException, zipfile.BadZipFile) as error:
        # Um .xlsx é um zip: conteúdo que não seja zip nunca é uma planilha.
        raise ExtractionError(
            f"'{path.name}' não é uma planilha Excel válida. "
            "Formatos aceitos: .xlsx e .xlsm. Converta arquivos .xls antes de usar."
        ) from error
    except (OSError, KeyError, ValueError) as error:
        raise ExtractionError(
            f"Não foi possível abrir '{path.name}': o arquivo parece corrompido."
        ) from error


def _select_sheet(
    workbook: Any, sheet_names: tuple[str, ...], sheet: str | None, path: Path
) -> Any:
    """Seleciona a aba pedida ou a primeira, com erro claro quando não existe."""
    if not sheet_names:
        raise EmptySourceError(f"A planilha '{path.name}' não possui abas.")
    if sheet is None:
        return workbook[sheet_names[0]]
    if sheet not in sheet_names:
        available = ", ".join(sheet_names)
        raise SheetNotFoundError(
            f"A aba '{sheet}' não existe em '{path.name}'. "
            f"Abas disponíveis: {available}."
        )
    return workbook[sheet]


def _iter_rows(worksheet: Any) -> Iterator[tuple[int, list[str | None]]]:
    """Itera ``(número da linha na planilha, células já em texto)``."""
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        yield row_number, [coerce_cell(cell) for cell in row]


def _read_header(
    rows: Iterator[tuple[int, list[str | None]]],
    header_row: int,
    path: Path,
    sheet_title: str,
) -> tuple[tuple[str, ...], tuple[str, ...]]:
    """Avança até ``header_row`` (ignorando títulos acima) e normaliza o cabeçalho."""
    for row_number, values in rows:
        if row_number < header_row:
            continue
        header, warnings = normalize_headers(values)
        if not header or is_blank_row(values):
            raise EmptySourceError(
                f"A linha {header_row} da aba '{sheet_title}' em '{path.name}' "
                "está vazia e não pode ser usada como cabeçalho."
            )
        return header, warnings

    raise EmptySourceError(
        f"A aba '{sheet_title}' de '{path.name}' não possui a linha de "
        f"cabeçalho {header_row}."
    )


def _read_records(
    rows: Iterator[tuple[int, list[str | None]]],
    header: tuple[str, ...],
    sheet_title: str,
) -> Iterator[SourceRecord]:
    """Converte as linhas restantes em registros, ignorando linhas em branco."""
    for row_number, values in rows:
        if is_blank_row(values):
            continue
        yield SourceRecord(
            row_number=row_number,
            values=build_row_mapping(header, values),
            sheet=sheet_title,
        )


__all__ = ["SUPPORTED_SUFFIXES", "ExcelExtractor"]
