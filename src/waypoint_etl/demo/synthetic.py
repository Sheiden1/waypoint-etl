"""Registros sintéticos de clientes no estilo de um ERP legado.

Gera dados propositalmente "sujos" para exercitar o pipeline de limpeza,
validação e deduplicação: máscaras variadas, datas em três formatos, duplicatas,
documentos e e-mails inválidos, campos vazios e ruído textual.

Os cabeçalhos usam nomes em português no estilo legado, compatíveis com o
exemplo de mapeamento De/Para do CLAUDE.md.
"""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from random import Random

from openpyxl import Workbook

from .documents import generate_cnpj, generate_cpf

# Data fixa gravada nos metadados dos arquivos gerados, para que o conteúdo não
# dependa do momento da execução. Os containers .xlsx/.docx ainda gravam a data
# de modificação de cada entrada do ZIP, por isso os binários não são
# versionados (ver .gitignore).
FIXTURE_TIMESTAMP = datetime(2024, 1, 1, 12, 0, 0)

# Cabeçalhos legados (origem). Não confundir com o schema canônico de destino.
CSV_HEADERS = [
    "Código",
    "Nome Cliente",
    "CPF_CNPJ",
    "Correio Eletrônico",
    "Fone Principal",
    "CEP",
    "Cidade",
    "UF",
    "Data Cadastro",
    "Observação Interna Antiga",
]

_FIRST_NAMES = [
    "Ana", "Bruno", "Carla", "Diego", "Elaine", "Fábio", "Gabriela",
    "Heitor", "Isabela", "João", "Larissa", "Marcos", "Natália",
    "Otávio", "Priscila", "Rafael", "Sônia", "Thiago", "Vanessa", "William",
]
_LAST_NAMES = [
    "Silva", "Souza", "Oliveira", "Santos", "Pereira", "Costa", "Almeida",
    "Nununes", "Carvalho", "Ribeiro", "Gomes", "Martins", "Araújo", "Barbosa",
]
_COMPANY_SUFFIXES = ["ME", "LTDA", "EIRELI", "S/A"]
_CITIES = [
    ("São Paulo", "SP"), ("Rio de Janeiro", "RJ"), ("Belo Horizonte", "MG"),
    ("Curitiba", "PR"), ("Porto Alegre", "RS"), ("Salvador", "BA"),
    ("Recife", "PE"), ("Fortaleza", "CE"), ("Manaus", "AM"), ("Goiânia", "GO"),
]
_DATE_FORMATS = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"]

# Cabeçalhos legados da aba de contatos da planilha de demonstração.
CONTACT_HEADERS = [
    "Código Cliente",
    "CPF_CNPJ Cliente",
    "Nome Contato",
    "Cargo",
    "Correio Eletrônico",
    "Fone",
]

# A aba de clientes começa com uma linha de título, como em exportações reais:
# o cabeçalho fica na linha 2 (ver ``header_row`` no template De/Para).
CUSTOMERS_SHEET_TITLE = "Relatório de Clientes - ERP Legado (export)"
CUSTOMERS_SHEET_NAME = "Clientes"
CONTACTS_SHEET_NAME = "Contatos"
CUSTOMERS_HEADER_ROW = 2

_ROLES = ["Comprador", "Financeiro", "Diretor", "Gerente", "Sócio"]


def _mask_cpf(digits: str, rng: Random) -> str:
    """Aplica (ou não) máscara a um CPF, para variar os formatos de origem."""
    if rng.random() < 0.5:
        return f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}"
    return digits


def _mask_cnpj(digits: str, rng: Random) -> str:
    """Aplica (ou não) máscara a um CNPJ."""
    if rng.random() < 0.5:
        return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:]}"
    return digits


def _phone(rng: Random) -> str:
    """Gera um telefone brasileiro com máscara variada."""
    ddd = rng.randint(11, 99)
    number = rng.randint(90000_0000, 99999_9999)
    style = rng.randint(0, 2)
    if style == 0:
        return f"({ddd}) {str(number)[:5]}-{str(number)[5:]}"
    if style == 1:
        return f"+55 {ddd} {str(number)[:5]}-{str(number)[5:]}"
    return f"{ddd}{number}"


def _format_date(rng: Random) -> str:
    """Gera uma data de cadastro em um dos três formatos suportados."""
    from datetime import date, timedelta

    start = date(2018, 1, 1)
    day = start + timedelta(days=rng.randint(0, 2500))
    return day.strftime(rng.choice(_DATE_FORMATS))


def generate_customer_rows(
    *, seed: int = 20240101, count: int = 50
) -> list[dict[str, str]]:
    """Gera registros de clientes sintéticos e determinísticos.

    Inclui, além dos ``count`` clientes principais: 5 duplicatas, 5 documentos
    inválidos, 3 e-mails inválidos, campos vazios e ruído textual.
    """
    rng = Random(seed)
    rows: list[dict[str, str]] = []

    for i in range(count):
        is_company = rng.random() < 0.35
        if is_company:
            name = f"{rng.choice(_LAST_NAMES)} {rng.choice(_COMPANY_SUFFIXES)}"
            document = _mask_cnpj(generate_cnpj(rng), rng)
        else:
            name = f"{rng.choice(_FIRST_NAMES)} {rng.choice(_LAST_NAMES)}"
            document = _mask_cpf(generate_cpf(rng), rng)

        city, uf = rng.choice(_CITIES)
        slug = name.lower().replace(" ", ".").replace("/", "")
        email = f"{slug}@exemplo.com.br"

        rows.append(
            {
                "Código": f"ERP-{1000 + i}",
                "Nome Cliente": name,
                "CPF_CNPJ": document,
                "Correio Eletrônico": email,
                "Fone Principal": _phone(rng),
                "CEP": f"{rng.randint(1000, 99999):05d}-{rng.randint(0, 999):03d}",
                "Cidade": city,
                "UF": uf,
                "Data Cadastro": _format_date(rng),
                "Observação Interna Antiga": rng.choice(
                    ["", "cliente antigo", "revisar", "N/A"]
                ),
            }
        )

    _inject_noise(rows, rng)
    _inject_invalid_documents(rows, rng)
    _inject_invalid_emails(rows, rng)
    _inject_empty_fields(rows, rng)
    _append_duplicates(rows, rng)
    return rows


def _inject_noise(rows: list[dict[str, str]], rng: Random) -> None:
    """Adiciona espaços extras e caracteres especiais a alguns nomes."""
    for index in rng.sample(range(len(rows)), k=min(6, len(rows))):
        name = rows[index]["Nome Cliente"]
        rows[index]["Nome Cliente"] = f"  {name.upper()}   "


def _inject_invalid_documents(rows: list[dict[str, str]], rng: Random) -> None:
    """Corrompe 5 documentos para exercitar a rejeição por dígito verificador."""
    for index in rng.sample(range(len(rows)), k=min(5, len(rows))):
        rows[index]["CPF_CNPJ"] = "111.111.111-11"


def _inject_invalid_emails(rows: list[dict[str, str]], rng: Random) -> None:
    """Corrompe 3 e-mails para exercitar a validação de e-mail."""
    broken = ["sem-arroba.com", "usuario@", "@dominio.com"]
    for offset, index in enumerate(
        rng.sample(range(len(rows)), k=min(3, len(rows)))
    ):
        rows[index]["Correio Eletrônico"] = broken[offset % len(broken)]


def _inject_empty_fields(rows: list[dict[str, str]], rng: Random) -> None:
    """Esvazia campos opcionais em alguns registros e insere marcadores nulos."""
    for index in rng.sample(range(len(rows)), k=min(5, len(rows))):
        rows[index]["Fone Principal"] = rng.choice(["", "-", "N/A"])
    for index in rng.sample(range(len(rows)), k=min(4, len(rows))):
        rows[index]["CEP"] = ""


def _append_duplicates(rows: list[dict[str, str]], rng: Random) -> None:
    """Acrescenta 5 duplicatas (mesmo documento, com pequenas variações)."""
    for source in rows[:5]:
        duplicate = dict(source)
        duplicate["Código"] = f"{source['Código']}-DUP"
        duplicate["Nome Cliente"] = source["Nome Cliente"].strip().title()
        duplicate["Correio Eletrônico"] = source["Correio Eletrônico"].upper()
        rows.append(duplicate)


def generate_contact_rows(
    customer_rows: list[dict[str, str]], *, seed: int = 20240101, count: int = 12
) -> list[dict[str, str]]:
    """Gera contatos sintéticos vinculados aos clientes informados.

    Inclui casos sem e-mail e sem telefone para exercitar a regra "pelo menos
    e-mail ou telefone" da validação de contatos (seção 13).
    """
    rng = Random(seed + 1)
    rows: list[dict[str, str]] = []

    for index, customer in enumerate(customer_rows[:count]):
        name = f"{rng.choice(_FIRST_NAMES)} {rng.choice(_LAST_NAMES)}"
        slug = name.lower().replace(" ", ".")
        rows.append(
            {
                "Código Cliente": customer["Código"],
                "CPF_CNPJ Cliente": customer["CPF_CNPJ"],
                "Nome Contato": name,
                "Cargo": rng.choice(_ROLES),
                # Um contato a cada quatro fica sem e-mail e outro sem telefone.
                "Correio Eletrônico": (
                    "" if index % 4 == 0 else f"{slug}@exemplo.com.br"
                ),
                "Fone": "" if index % 4 == 1 else _phone(rng),
            }
        )

    return rows


def write_customers_csv(path: Path, *, seed: int = 20240101, count: int = 50) -> Path:
    """Escreve o CSV de clientes legados em ``path`` e retorna o caminho."""
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = generate_customer_rows(seed=seed, count=count)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=CSV_HEADERS)
        writer.writeheader()
        writer.writerows(rows)
    return path


def write_legacy_xlsx(path: Path, *, seed: int = 20240101, count: int = 50) -> Path:
    """Escreve a planilha legada com duas abas (``Clientes`` e ``Contatos``).

    A aba de clientes traz uma linha de título antes do cabeçalho, exercitando o
    ``header_row`` configurável dos templates De/Para.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    customer_rows = generate_customer_rows(seed=seed, count=count)
    contact_rows = generate_contact_rows(customer_rows, seed=seed)

    workbook = Workbook()
    workbook.properties.created = FIXTURE_TIMESTAMP
    workbook.properties.modified = FIXTURE_TIMESTAMP
    default_sheet = workbook.active
    if default_sheet is not None:
        workbook.remove(default_sheet)

    customers = workbook.create_sheet(CUSTOMERS_SHEET_NAME)
    customers.append([CUSTOMERS_SHEET_TITLE])
    customers.append(CSV_HEADERS)
    for row in customer_rows:
        customers.append([row[header] for header in CSV_HEADERS])

    contacts = workbook.create_sheet(CONTACTS_SHEET_NAME)
    contacts.append(CONTACT_HEADERS)
    for contact in contact_rows:
        contacts.append([contact[header] for header in CONTACT_HEADERS])

    workbook.save(path)
    workbook.close()
    return path
